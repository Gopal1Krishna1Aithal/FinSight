import os
import sys
import re

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

import pandas as pd
from dotenv import load_dotenv

from core.extractors.hdfc_pdf import HDFCPDFExtractor
from core.extractors.universal_pdf import UniversalPDFExtractor
from core.extractors.extraction_validator import ExtractionValidator
from core.processors.cleaner import HDFCDataCleaner
from core.processors.sanitizer import DataSanitizer
from core.ai_services.coa_mapper import CoAMapper, CONFIDENCE_THRESHOLD

from core.db.session import init_db
from core.db.operations import upsert_transactions
from core.ai_services.insights_generator import InsightsGenerator
from core.processors.analysis_engine import FrontendDataEngine
import argparse

load_dotenv()

# ---------------------------------------------------------------------------
# CLI args
# ---------------------------------------------------------------------------

parser = argparse.ArgumentParser(description="Process a business bank statement PDF.")
parser.add_argument(
    "pdf",
    nargs="?",
    default=os.path.join("data", "input", "pdfs"),
    help="Path to a single PDF or a folder of PDFs",
)
parser.add_argument(
    "--extractor",
    choices=["universal", "hdfc"],
    default="hdfc",
    help="Extraction engine: 'hdfc' (default) for HDFC PDF parser; 'universal' for Gemini AI.",
)
args = parser.parse_args()

PDF_PATH     = args.pdf
OUT_DIR      = os.path.join("data", "output")
EXCEL_PATH   = os.path.join(OUT_DIR, "clean_statement.xlsx")
TALLY_PATH   = os.path.join(OUT_DIR, "tally_import.csv")
TALLY_XML    = os.path.join(OUT_DIR, "tally_import.xml")
INSIGHTS_PATH = os.path.join(OUT_DIR, "financial_insights.md")

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

# Quarter → month labels for nicer sheet names
_QUARTER_LABELS = {
    "Q1": "Apr–Jun",
    "Q2": "Jul–Sep",
    "Q3": "Oct–Dec",
    "Q4": "Jan–Mar",
}

# Financial year → start year for short label
_FY_LABELS = {
    "FY2324": "23",
    "FY2425": "24",
    "FY2526": "25",
}


def _infer_period_label(filename: str, df: pd.DataFrame = None) -> tuple[str, str]:
    """
    Return (period_label, sheet_name) from a filename.
    Pattern: *_Q1_FY2324_* → ("Q1 FY2324", "Q1 Apr–Jun 23")
    Falls back to date range if no match.
    """
    stem = os.path.basename(filename)
    m = re.search(r"(Q[1-4])[\s_-]*(FY\d{4})", stem, re.IGNORECASE)
    if m:
        q   = m.group(1).upper()
        fy  = m.group(2).upper()
        qm  = _QUARTER_LABELS.get(q, "?")
        yy  = _FY_LABELS.get(fy, fy[-2:])
        period_label = f"{q} {fy}"
        sheet_name   = f"{q} {qm} {yy}"
        return period_label, sheet_name

    # Fallback: derive from data date range
    if df is not None and not df.empty and "Date" in df.columns:
        d_min = df["Date"].min().strftime("%b %Y")
        d_max = df["Date"].max().strftime("%b %Y")
        label = f"{d_min}–{d_max}"
        return label, label[:18]

    return "FY2324", "Transactions"


def _extract_single_pdf(pdf_file: str) -> list[dict]:
    if args.extractor == "hdfc":
        return HDFCPDFExtractor(pdf_file).extract() or []

    gemini_key = os.getenv("GEMINI_API_KEY")
    if not gemini_key:
        print("      [!] GEMINI_API_KEY not set — falling back to HDFCPDFExtractor.")
        return HDFCPDFExtractor(pdf_file).extract() or []

    try:
        data = UniversalPDFExtractor(pdf_file, api_key=gemini_key).extract()
    except Exception as e:
        print(f"      [Universal] Error: {e}")
        data = None

    if not data:
        print("      [Universal] ⚠  0 rows — falling back to HDFCPDFExtractor...")
        return HDFCPDFExtractor(pdf_file).extract() or []

    return data


def validate_balances(df: pd.DataFrame) -> bool:
    TOLERANCE = 0.01
    prev_balance = df.iloc[0]["Balance"] + df.iloc[0]["Debit"] - df.iloc[0]["Credit"]
    for idx, row in df.iterrows():
        expected = prev_balance - row["Debit"] + row["Credit"]
        actual   = row["Balance"]
        if abs(expected - actual) > TOLERANCE:
            print(
                f"\n      [VALIDATOR] ❌  Mismatch on row {idx} "
                f"({row['Date'].strftime('%d/%m/%Y')}):\n"
                f"        Expected : ₹{expected:,.2f}\n"
                f"        Actual   : ₹{actual:,.2f}"
            )
            return False
        prev_balance = actual
    return True


# ---------------------------------------------------------------------------
# Cross-quarter continuity check
# ---------------------------------------------------------------------------

def check_continuity(quarters: list[tuple]) -> list[dict]:
    """
    quarters: list of (period_label, df) sorted chronologically.
    Returns a list of warning dicts.
    """
    warnings = []
    TOLERANCE = 1.0  # ₹1 tolerance for rounding differences

    print("\n  ┌─ Cross-Quarter Balance Continuity ─────────────────────────┐")
    for i in range(len(quarters) - 1):
        label_a, df_a = quarters[i]
        label_b, df_b = quarters[i + 1]

        closing  = float(df_a.iloc[-1]["Balance"])
        opening  = float(df_b.iloc[0]["Balance"] + df_b.iloc[0]["Debit"] - df_b.iloc[0]["Credit"])
        diff     = abs(closing - opening)
        ok       = diff <= TOLERANCE
        symbol   = "✅" if ok else "⚠️ "
        print(
            f"  │  {label_a} closing → {label_b} opening : "
            f"₹{closing:>14,.2f} → ₹{opening:>14,.2f}  {symbol}"
        )
        if not ok:
            warnings.append({
                "from": label_a,
                "to":   label_b,
                "closing_balance":  round(closing, 2),
                "opening_balance":  round(opening, 2),
                "difference":       round(diff, 2),
                "message": (
                    f"{label_a} closing balance (₹{closing:,.2f}) does not match "
                    f"{label_b} opening balance (₹{opening:,.2f}) — "
                    f"gap of ₹{diff:,.2f}"
                ),
            })
    print("  └─────────────────────────────────────────────────────────────┘")
    return warnings


# ---------------------------------------------------------------------------
# Excel — 7-sheet workbook
# ---------------------------------------------------------------------------

def _save_excel_multiperiod(quarters: list[tuple], combined_df: pd.DataFrame, path: str) -> None:
    """
    Writes a 7-sheet Excel workbook:
      Q1 Apr–Jun 23 … Q4 Jan–Mar 24  (per-quarter sheets)
      All Transactions               (combined + Period column)
      Period Summary                  (Q1→Q4 trend table)
      Annual Summary                  (full-year category totals)
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        print("      [!] openpyxl not installed.")
        return

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    quarter_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    green_fill   = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill     = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    review_fill  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    title_font   = Font(bold=True, size=13, color="1F4E79")
    bold_font    = Font(bold=True, size=10)
    green_font   = Font(bold=True, color="375623", size=10)
    red_font     = Font(bold=True, color="C00000", size=10)

    txn_cols = ["Date", "Narration", "Clean_Description", "CoA_Category",
                "Confidence_Score", "Reasoning", "Review_Required", "Debit", "Credit", "Balance"]
    col_widths = {"A": 12, "B": 45, "C": 28, "D": 26, "E": 16,
                  "F": 40, "G": 16, "H": 14, "I": 14, "J": 14}

    def _style_txn_sheet(ws, df_sheet):
        ws.freeze_panes = "A2"
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30
        flag_col_idx = txn_cols.index("Review_Required") + 1
        for row_idx, rv in enumerate(df_sheet["Review_Required"], start=2):
            if rv:
                for ci in range(1, len(txn_cols) + 1):
                    ws.cell(row=row_idx, column=ci).fill = review_fill
                fc = ws.cell(row=row_idx, column=flag_col_idx)
                fc.font = Font(bold=True, color="C00000")
                fc.alignment = Alignment(horizontal="center")

    def _prep_txn_df(df):
        out = df[["Date","Narration","Clean_Description","CoA_Category",
                   "Confidence_Score","Reasoning","Debit","Credit","Balance"]].copy()
        out["Date"] = df["Date"].dt.strftime("%d/%m/%Y")
        out["Review_Required"] = (
            (df["Confidence_Score"] < CONFIDENCE_THRESHOLD) |
            (df["CoA_Category"] == "Uncategorized")
        )
        return out[txn_cols]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:

        # ── Per-quarter sheets ─────────────────────────────────────────
        for period_label, sheet_name, df_q in quarters:
            out_q = _prep_txn_df(df_q)
            out_q.to_excel(writer, sheet_name=sheet_name[:31], index=False)
            ws = writer.sheets[sheet_name[:31]]
            _style_txn_sheet(ws, out_q)

        # ── All Transactions sheet ─────────────────────────────────────
        all_df = combined_df.copy()
        all_df["Date"] = all_df["Date"].dt.strftime("%d/%m/%Y")
        all_df["Review_Required"] = (
            (combined_df["Confidence_Score"] < CONFIDENCE_THRESHOLD) |
            (combined_df["CoA_Category"] == "Uncategorized")
        )
        all_cols = ["Date","Period","Narration","Clean_Description","CoA_Category",
                    "Confidence_Score","Reasoning","Review_Required","Debit","Credit","Balance"]
        all_df = all_df[all_cols]
        all_df.to_excel(writer, sheet_name="All Transactions", index=False)
        ws_all = writer.sheets["All Transactions"]
        ws_all.freeze_panes = "A2"
        all_col_widths = {"A":12,"B":14,"C":45,"D":28,"E":26,"F":16,"G":40,"H":16,"I":14,"J":14,"K":14}
        for col, w in all_col_widths.items():
            ws_all.column_dimensions[col].width = w
        for cell in ws_all[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_all.row_dimensions[1].height = 30

        # ── Period Summary sheet ───────────────────────────────────────
        ps_rows = []
        for period_label, sheet_name, df_q in quarters:
            opening = float(df_q.iloc[0]["Balance"] + df_q.iloc[0]["Debit"] - df_q.iloc[0]["Credit"])
            closing = float(df_q.iloc[-1]["Balance"])
            ps_rows.append({
                "Period":          period_label,
                "Date Range":      f"{df_q['Date'].min().strftime('%d %b %Y')} – {df_q['Date'].max().strftime('%d %b %Y')}",
                "Transactions":    len(df_q),
                "Total Inflow":    round(df_q["Credit"].sum(), 2),
                "Total Outflow":   round(df_q["Debit"].sum(), 2),
                "Net Cash Flow":   round(df_q["Credit"].sum() - df_q["Debit"].sum(), 2),
                "Opening Balance": round(opening, 2),
                "Closing Balance": round(closing, 2),
            })

        ps_df = pd.DataFrame(ps_rows)
        ps_df.to_excel(writer, sheet_name="Period Summary", index=False, startrow=1)
        ws_ps = writer.sheets["Period Summary"]
        ws_ps["A1"] = "QUARTERLY PERIOD SUMMARY — FY2324"
        ws_ps["A1"].font = title_font
        ps_col_widths = {"A":14,"B":28,"C":14,"D":16,"E":16,"F":16,"G":18,"H":18}
        for col, w in ps_col_widths.items():
            ws_ps.column_dimensions[col].width = w
        for cell in ws_ps[2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_ps.row_dimensions[2].height = 28
        # Colour net flow cells
        net_col_idx = list(ps_df.columns).index("Net Cash Flow") + 1
        for r_idx, val in enumerate(ps_df["Net Cash Flow"], start=3):
            cell = ws_ps.cell(row=r_idx, column=net_col_idx)
            cell.fill = green_fill if val >= 0 else red_fill
            cell.font = Font(bold=True, color="375623" if val >= 0 else "C00000", size=10)

        # Bold totals row at bottom
        total_row = len(ps_rows) + 3
        totals = {
            "Period": "── ANNUAL TOTAL ──",
            "Transactions": combined_df.shape[0],
            "Total Inflow": round(combined_df["Credit"].sum(), 2),
            "Total Outflow": round(combined_df["Debit"].sum(), 2),
            "Net Cash Flow": round(combined_df["Credit"].sum() - combined_df["Debit"].sum(), 2),
        }
        for c_idx, col_name in enumerate(ps_df.columns, start=1):
            cell = ws_ps.cell(row=total_row, column=c_idx)
            cell.value = totals.get(col_name, "")
            cell.font  = bold_font
            total_fill_c = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
            cell.fill  = total_fill_c

        # ── Annual Summary sheet ───────────────────────────────────────
        sum_rows = []
        for cat in sorted(combined_df["CoA_Category"].unique()):
            cat_df = combined_df[combined_df["CoA_Category"] == cat]
            sum_rows.append({
                "Category":     cat,
                "Txn Count":    len(cat_df),
                "Total Debit":  round(cat_df["Debit"].sum(), 2),
                "Total Credit": round(cat_df["Credit"].sum(), 2),
                "Net":          round(cat_df["Credit"].sum() - cat_df["Debit"].sum(), 2),
            })
        sum_df = pd.DataFrame(sum_rows)

        opening_bal = float(combined_df.sort_values("Date").iloc[0]["Balance"] +
                           combined_df.sort_values("Date").iloc[0]["Debit"] -
                           combined_df.sort_values("Date").iloc[0]["Credit"])
        closing_bal = float(combined_df.sort_values("Date").iloc[-1]["Balance"])

        sum_df.to_excel(writer, sheet_name="Annual Summary", index=False, startrow=6)
        ws_ann = writer.sheets["Annual Summary"]
        ws_ann["A1"] = "ANNUAL FINANCIAL SUMMARY — FY 2023–24"
        ws_ann["A1"].font = title_font
        for col, w in {"A":28,"B":12,"C":16,"D":16,"E":14}.items():
            ws_ann.column_dimensions[col].width = w

        kv = [
            ("Opening Balance", f"₹{opening_bal:,.2f}"),
            ("Closing Balance", f"₹{closing_bal:,.2f}"),
            ("Total Inflow",    f"₹{combined_df['Credit'].sum():,.2f}"),
            ("Total Outflow",   f"₹{combined_df['Debit'].sum():,.2f}"),
            ("Net Cash Flow",   f"₹{combined_df['Credit'].sum() - combined_df['Debit'].sum():,.2f}"),
        ]
        for i, (lbl, val) in enumerate(kv, start=2):
            ws_ann.cell(row=i, column=1).value = lbl
            ws_ann.cell(row=i, column=1).font  = bold_font
            ws_ann.cell(row=i, column=2).value = val
            net = combined_df["Credit"].sum() - combined_df["Debit"].sum()
            ws_ann.cell(row=i, column=2).font = (
                green_font if lbl == "Net Cash Flow" and net >= 0
                else red_font  if lbl == "Net Cash Flow"
                else Font(size=10)
            )
        for cell in ws_ann[7]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    review_count = int(
        ((combined_df["Confidence_Score"] < CONFIDENCE_THRESHOLD) |
         (combined_df["CoA_Category"] == "Uncategorized")).sum()
    )
    print(f"      ✅  Excel (7 sheets) → {path}  ({review_count} rows flagged)")


def _save_tally_csv(df: pd.DataFrame, path: str) -> None:
    tally = df[["Date", "Clean_Description", "CoA_Category", "Debit", "Credit"]].copy()
    tally["Date"] = df["Date"].dt.strftime("%d/%m/%Y")
    tally["Voucher_Type"] = tally.apply(lambda r: "Payment" if r["Debit"] > 0 else "Receipt", axis=1)
    tally["Amount"] = tally.apply(lambda r: r["Debit"] if r["Debit"] > 0 else r["Credit"], axis=1)
    tally = tally.rename(columns={"Clean_Description": "Ledger_Name"})
    if "Period" in df.columns:
        tally["Period"] = df["Period"]
    tally[["Date", "Voucher_Type", "Ledger_Name", "CoA_Category", "Amount"]].to_csv(path, index=False)
    print(f"      ✅  Tally CSV → {path}")


def _save_tally_xml(df: pd.DataFrame, path: str) -> None:
    import xml.etree.ElementTree as ET
    from xml.dom import minidom

    envelope  = ET.Element("ENVELOPE")
    header    = ET.SubElement(envelope, "HEADER")
    ET.SubElement(header, "TALLYREQUEST").text = "Import Data"
    body      = ET.SubElement(envelope, "BODY")
    importdata = ET.SubElement(body, "IMPORTDATA")
    reqdesc   = ET.SubElement(importdata, "REQUESTDESC")
    ET.SubElement(reqdesc, "REPORTNAME").text = "Vouchers"
    staticvars = ET.SubElement(reqdesc, "STATICVARIABLES")
    ET.SubElement(staticvars, "SVCURRENTCOMPANY").text = "My Company"
    reqdata   = ET.SubElement(importdata, "REQUESTDATA")

    for _, row in df.iterrows():
        if row["Debit"] == 0 and row["Credit"] == 0:
            continue
        tmsg    = ET.SubElement(reqdata, "TALLYMESSAGE", {"xmlns:UDF": "TallyUDF"})
        vchtype = "Payment" if row["Debit"] > 0 else "Receipt"
        amt     = row["Debit"] if row["Debit"] > 0 else row["Credit"]
        voucher = ET.SubElement(tmsg, "VOUCHER", {"VCHTYPE": vchtype, "ACTION": "Create"})
        ET.SubElement(voucher, "DATE").text = row["Date"].strftime("%Y%m%d")
        ET.SubElement(voucher, "VOUCHERTYPENAME").text = vchtype
        ET.SubElement(voucher, "NARRATION").text = str(row.get("Clean_Description", ""))
        party = ET.SubElement(voucher, "ALLLEDGERENTRIES.LIST")
        ET.SubElement(party, "LEDGERNAME").text = str(row.get("Clean_Description", ""))
        ET.SubElement(party, "ISDEEMEDPOSITIVE").text = "Yes" if vchtype == "Payment" else "No"
        ET.SubElement(party, "AMOUNT").text = f"-{amt}" if vchtype == "Payment" else f"{amt}"
        bank = ET.SubElement(voucher, "ALLLEDGERENTRIES.LIST")
        ET.SubElement(bank, "LEDGERNAME").text = "HDFC Bank"
        ET.SubElement(bank, "ISDEEMEDPOSITIVE").text = "No" if vchtype == "Payment" else "Yes"
        ET.SubElement(bank, "AMOUNT").text = f"{amt}" if vchtype == "Payment" else f"-{amt}"

    xml_str = minidom.parseString(ET.tostring(envelope)).toprettyxml(indent="  ")
    with open(path, "w", encoding="utf-8") as f:
        f.write(xml_str)
    print(f"      ✅  Tally XML → {path}")


# ---------------------------------------------------------------------------
# Pipeline
# ---------------------------------------------------------------------------

def run_pipeline() -> None:
    global PDF_PATH
    os.makedirs(OUT_DIR, exist_ok=True)

    print("\n[0/7] Initializing database (applying additive migrations)...")
    init_db()

    # ── Step 1: Collect PDF files ─────────────────────────────────────────
    print(f"\n[1/7] Locating PDF files in '{PDF_PATH}'...")

    if os.path.isfile(PDF_PATH):
        pdf_files = [PDF_PATH]
    elif os.path.isdir(PDF_PATH):
        pdf_files = sorted([
            os.path.join(PDF_PATH, f)
            for f in os.listdir(PDF_PATH)
            if f.lower().endswith(".pdf")
        ])
        if not pdf_files:
            # Fallback to images folder
            img_dir = os.path.join("data", "input", "images")
            img_exts = {".jpg", ".jpeg", ".png", ".heic"}
            if os.path.exists(img_dir):
                img_files = [os.path.join(img_dir, f) for f in os.listdir(img_dir)
                             if os.path.splitext(f.lower())[1] in img_exts]
                if img_files:
                    print(f"      [!] No PDFs found. Falling back to image OCR: {img_dir}")
                    from core.extractors.image_ocr import ImageOCRExtractor
                    raw = ImageOCRExtractor(image_paths=img_files).extract() or []
                    if not raw:
                        print("      [!] OCR returned no data. Aborting.")
                        sys.exit(1)
                    clean_df  = HDFCDataCleaner(raw).clean()
                    safe_df   = DataSanitizer(clean_df).scrub_pii()
                    safe_df["CoA_Category"]     = "Uncategorized"
                    safe_df["Confidence_Score"] = 0
                    safe_df["Reasoning"]        = "Image OCR — no CoA."
                    _save_tally_csv(safe_df, TALLY_PATH)
                    _save_tally_xml(safe_df, TALLY_XML)
                    upsert_transactions(safe_df)
                    return
            print(f"      [!] No PDF files found in '{PDF_PATH}'. Aborting.")
            sys.exit(1)
    else:
        print(f"      [!] Path not found: '{PDF_PATH}'. Aborting.")
        sys.exit(1)

    print(f"      Found {len(pdf_files)} PDF file(s): {[os.path.basename(p) for p in pdf_files]}")
    multi_mode = len(pdf_files) > 1

    # ── Step 1.5 → Validate + Clean + CoA per PDF ───────────────────────
    groq_api_key = os.getenv("GROQ_API_KEY")
    quarters: list[tuple] = []   # (period_label, sheet_name, clean_safe_df)

    for pdf_file in pdf_files:
        fname = os.path.basename(pdf_file)
        print(f"\n  ── Processing: {fname}")

        # Extract
        raw_data = _extract_single_pdf(pdf_file)
        if not raw_data:
            print(f"      [!] Extraction returned no data for {fname}. Skipping.")
            continue
        print(f"      → {len(raw_data)} raw rows extracted.")

        # Validate extraction
        val = ExtractionValidator(raw_data).validate()
        print(val.report())
        if not val.passed:
            print(f"      [⚠] Extraction validation failed for {fname}. Skipping.")
            continue

        # Clean
        clean_df = HDFCDataCleaner(raw_data).clean()
        if clean_df.empty:
            print(f"      [!] Clean step produced 0 rows for {fname}. Skipping.")
            continue

        # Sanitize PII
        safe_df = DataSanitizer(clean_df).scrub_pii()

        # CoA categorize
        if groq_api_key:
            safe_df = CoAMapper(api_key=groq_api_key).map(safe_df)
        else:
            safe_df["CoA_Category"]     = "Uncategorized"
            safe_df["Confidence_Score"] = 0
            safe_df["Reasoning"]        = "GROQ_API_KEY not set."

        # Infer period label from filename
        period_label, sheet_name = _infer_period_label(pdf_file, safe_df)
        safe_df["Period"]       = period_label
        safe_df["Source_File"]  = fname

        quarters.append((period_label, sheet_name, safe_df))
        print(f"      ✅  {period_label} ({sheet_name}): {len(safe_df)} rows tagged.")

    if not quarters:
        print("\n[!] No valid data extracted from any PDF. Aborting.")
        sys.exit(1)

    # Sort quarters chronologically by first date in each df
    quarters.sort(key=lambda t: t[2]["Date"].min())

    # ── Step 1.6: Cross-quarter continuity check ─────────────────────────
    print("\n[1.6/7] Cross-quarter balance continuity check...")
    continuity_pairs = [(q[0], q[2]) for q in quarters]
    continuity_warnings = check_continuity(continuity_pairs) if len(quarters) > 1 else []
    if not continuity_warnings:
        print("      ✅  All quarter balances chain correctly.")
    else:
        print(f"      ⚠️   {len(continuity_warnings)} continuity gap(s) detected (non-fatal).")

    # ── Step 5: Save outputs ──────────────────────────────────────────────
    print("\n[5/7] Writing output files and updating database...")

    # Combined DataFrame (with Period column)
    combined_df = pd.concat([q[2] for q in quarters], ignore_index=True)
    combined_df = combined_df.sort_values("Date").reset_index(drop=True)

    if multi_mode:
        _save_excel_multiperiod(quarters, combined_df, EXCEL_PATH)
    else:
        # Single-file path: use original simple Excel writer
        _save_excel_single(quarters[0][2], EXCEL_PATH)

    _save_tally_csv(combined_df, TALLY_PATH)
    _save_tally_xml(combined_df, TALLY_XML)

    # DB upsert per quarter
    total_new = 0
    for period_label, sheet_name, df_q in quarters:
        source_file  = df_q["Source_File"].iloc[0] if "Source_File" in df_q.columns else "UNKNOWN"
        new_rows = upsert_transactions(df_q, source_file=source_file, period_label=period_label)
        total_new += new_rows
        print(f"      DB ← {period_label}: {new_rows} new rows inserted.")
    print(f"      ✅  Database → {total_new} total new transactions.")

    # ── Step 6: Insights ─────────────────────────────────────────────────
    print("\n[6/7] Generating financial insights from full history...")
    try:
        InsightsGenerator().generate_insights(INSIGHTS_PATH)
    except Exception as e:
        print(f"      [Insights] Error: {e}")

    # ── Step 7: Frontend JSON ────────────────────────────────────────────
    print("\n[7/7] Generating frontend API JSON...")
    try:
        fe = FrontendDataEngine()
        fe.generate()
    except Exception as e:
        print(f"      [FrontendDataEngine] Error: {e}")

    # ── Terminal summary ─────────────────────────────────────────────────
    net = combined_df["Credit"].sum() - combined_df["Debit"].sum()
    print(f"\n{'─' * 62}")
    print(f"  Pipeline complete — {len(combined_df)} transactions across {len(quarters)} period(s)")
    print(f"  Period(s)       : {', '.join(q[0] for q in quarters)}")
    print(f"  Total Inflow    : ₹{combined_df['Credit'].sum():>14,.2f}")
    print(f"  Total Outflow   : ₹{combined_df['Debit'].sum():>14,.2f}")
    print(f"  Net Cash Flow   : ₹{net:>14,.2f}  {'▲' if net >= 0 else '▼'}")
    print(f"  Closing Balance : ₹{combined_df.iloc[-1]['Balance']:>14,.2f}")
    if continuity_warnings:
        print(f"\n  ⚠️  Continuity warnings:")
        for w in continuity_warnings:
            print(f"     • {w['message']}")
    print(f"\n  Output : {os.path.abspath(OUT_DIR)}")
    print(f"{'─' * 62}\n")


# ---------------------------------------------------------------------------
# Single-file Excel (unchanged logic, kept for backward compatibility)
# ---------------------------------------------------------------------------

def _save_excel_single(df: pd.DataFrame, path: str) -> None:
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        return

    out = df[["Date","Narration","Clean_Description","CoA_Category",
              "Confidence_Score","Reasoning","Debit","Credit","Balance"]].copy()
    out["Date"] = df["Date"].dt.strftime("%d/%m/%Y")
    out["Review_Required"] = (
        (df["Confidence_Score"] < CONFIDENCE_THRESHOLD) |
        (df["CoA_Category"] == "Uncategorized")
    )
    out = out[["Date","Narration","Clean_Description","CoA_Category",
               "Confidence_Score","Reasoning","Review_Required","Debit","Credit","Balance"]]

    summary_rows = []
    for cat in sorted(df["CoA_Category"].unique()):
        cdf = df[df["CoA_Category"] == cat]
        summary_rows.append({"Category": cat, "Txn Count": len(cdf),
                              "Total Debit": round(cdf["Debit"].sum(), 2),
                              "Total Credit": round(cdf["Credit"].sum(), 2),
                              "Net": round(cdf["Credit"].sum() - cdf["Debit"].sum(), 2)})
    sum_df = pd.DataFrame(summary_rows)

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        out.to_excel(writer, sheet_name="Transactions", index=False)
        ws = writer.sheets["Transactions"]
        ws.freeze_panes = "A2"
        for col, w in {"A":12,"B":45,"C":28,"D":26,"E":16,"F":40,"G":16,"H":14,"I":14,"J":14}.items():
            ws.column_dimensions[col].width = w
        for cell in ws[1]:
            cell.fill = header_fill; cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30

        sum_df.to_excel(writer, sheet_name="Summary", index=False, startrow=2)
        ws2 = writer.sheets["Summary"]
        ws2["A1"] = "Annual Summary"
        ws2["A1"].font = Font(bold=True, size=13, color="1F4E79")

    print(f"      ✅  Excel → {path}")


if __name__ == "__main__":
    run_pipeline()
