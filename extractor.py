import pdfplumber
import json
import re
import os
import time
import xml.etree.ElementTree as ET
from xml.dom import minidom
from collections import defaultdict
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font
from dotenv import load_dotenv
from groq import Groq

load_dotenv()

DATE_X_MAX      = 70
NARRATION_X_MIN = 71
NARRATION_X_MAX = 288
DEBIT_X_MIN     = 405
DEBIT_X_MAX     = 490
CREDIT_X_MIN    = 491
CREDIT_X_MAX    = 563
BALANCE_X_MIN   = 564


def is_date(s):
    return bool(re.match(r'^\d{2}/\d{2}/\d{2,4}$', s.strip()))


def clean_narration(s):
    s = re.sub(r'\b\d{10,}\b', '', s)
    s = re.sub(r'\b\d{6}[X*]+\d{4}\b', '', s)
    s = re.sub(r'\bS\s*DEBIT\b|\bS\s*CREDIT\b|\bEBIT\b', '', s, flags=re.IGNORECASE)
    return ' '.join(s.split()).strip()


def clean_amount(s):
    if not s or not str(s).strip():
        return 0.0
    try:
        return float(str(s).replace(',', '').strip())
    except:
        return 0.0


def parse_date(date_str):
    for fmt in ("%d/%m/%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except ValueError:
            continue
    return None


def month_key(date_str):
    dt = parse_date(date_str)
    return dt.strftime("%Y-%m") if dt else "0000-00"


# ════════════════════════════════════════════════════════════════════════════
# AGENT 1 & 2 — EXTRACTOR + VALIDATOR
# ════════════════════════════════════════════════════════════════════════════

def extract_hdfc_statement(pdf_path: str, password: str = ""):
    transactions = []
    try:
        with pdfplumber.open(pdf_path, password=password) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                words = page.extract_words(keep_blank_chars=False, x_tolerance=3, y_tolerance=3)
                if not words:
                    return {"status": "error", "message": f"Page {page_num} appears to be a scanned image."}

                rows = {}
                for w in words:
                    y = round(w['top'] / 3) * 3
                    rows.setdefault(y, []).append(w)

                current = None
                for y in sorted(rows.keys()):
                    row_words    = sorted(rows[y], key=lambda w: w['x0'])
                    date_text    = ' '.join(w['text'] for w in row_words if w['x0'] <= DATE_X_MAX)
                    narr_text    = ' '.join(w['text'] for w in row_words if NARRATION_X_MIN <= w['x0'] <= NARRATION_X_MAX)
                    debit_text   = next((w['text'] for w in row_words if DEBIT_X_MIN  <= w['x0'] <= DEBIT_X_MAX),  '')
                    credit_text  = next((w['text'] for w in row_words if CREDIT_X_MIN <= w['x0'] <= CREDIT_X_MAX), '')
                    balance_text = next((w['text'] for w in row_words if w['x0'] >= BALANCE_X_MIN), '')

                    if is_date(date_text):
                        if current:
                            transactions.append(current)
                        current = {
                            "Date": date_text, "Narration": narr_text,
                            "Debit": debit_text, "Credit": credit_text, "Balance": balance_text,
                        }
                    elif current and narr_text and not date_text:
                        current["Narration"] += " " + narr_text

                if current:
                    transactions.append(current)
                    current = None

    except Exception as e:
        return {"status": "error", "message": str(e)}

    if not transactions:
        return {"status": "error", "message": "No transactions found. Check if PDF is password protected or scanned."}

    for t in transactions:
        t["Narration"] = clean_narration(t["Narration"])

    prev_bal = None
    for t in transactions:
        d = clean_amount(t["Debit"])
        c = clean_amount(t["Credit"])
        b = clean_amount(t["Balance"])
        if prev_bal is not None and b != 0:
            expected = round(prev_bal - d + c, 2)
            t["ValidationFlag"] = "OK" if abs(expected - b) <= 1.0 else "BALANCE_MISMATCH"
        else:
            t["ValidationFlag"] = "OK"
        if b != 0:
            prev_bal = b

    return {"status": "success", "data": transactions}


# ════════════════════════════════════════════════════════════════════════════
# AGENT 3 — CATEGORIZER (Groq / Llama)
# ════════════════════════════════════════════════════════════════════════════

def categorize_with_groq(transactions: list) -> list:
    api_key = os.getenv("GROQ_API_KEY")
    if not api_key:
        print("Agent 3 Error: GROQ_API_KEY not found in .env")
        for t in transactions:
            t["CoA_Category"] = "Uncategorized / Flag for Review"
        return transactions

    client = Groq(api_key=api_key)
    categories_list = """
    - Payroll & Salaries
    - Software & Cloud Subscriptions
    - Office Rent & Utilities
    - Vendor Payments / Cost of Goods
    - Tax Payments (GST, TDS)
    - Bank Charges & Fees
    - Revenue / Inward Payment
    - Director Drawings / Personal
    - Travel & Transport
    - ATM Withdrawal
    - Credit Card Payment
    - Uncategorized / Flag for Review
    """
    chunk_size = 50

    for chunk_start in range(0, len(transactions), chunk_size):
        chunk = transactions[chunk_start: chunk_start + chunk_size]
        narration_map = {
            str(local_idx): {
                "text":   t["Narration"],
                "type":   "DEBIT (Expense)" if t["Debit"] else "CREDIT (Income)",
                "amount": t["Debit"] if t["Debit"] else t["Credit"],
            }
            for local_idx, t in enumerate(chunk)
        }
        system_prompt = f"""You are an expert Indian Chartered Accountant analyzing a Business Current Account.
Categorize each transaction into EXACTLY ONE of these categories:
{categories_list}

Input Data:
{json.dumps(narration_map, indent=2)}

IMPORTANT: Return a key for EVERY id. If unsure, use "Uncategorized / Flag for Review".
Reply ONLY in valid JSON mapping ID to Category.
Example: {{"0": "Vendor Payments / Cost of Goods", "1": "Bank Charges & Fees"}}"""

        print(f"Agent 3: Categorizing {chunk_start} to {chunk_start + len(chunk) - 1}...")
        category_mapping = {}

        for attempt in range(2):
            try:
                res = client.chat.completions.create(
                    messages=[
                        {"role": "system", "content": "You are a financial data categorization API. Output only JSON."},
                        {"role": "user",   "content": system_prompt},
                    ],
                    model="llama-3.3-70b-versatile",
                    response_format={"type": "json_object"},
                    temperature=0.1,
                )
                category_mapping = json.loads(res.choices[0].message.content)
                missing = [str(i) for i in range(len(chunk)) if str(i) not in category_mapping]
                if missing and attempt == 0:
                    print(f"  Groq missed {len(missing)} ids, retrying...")
                    time.sleep(1)
                    continue
                break
            except Exception as e:
                print(f"  Error attempt {attempt+1}: {e}")
                if attempt == 0:
                    time.sleep(2)

        for local_idx, t in enumerate(chunk):
            t["CoA_Category"] = category_mapping.get(str(local_idx), "Uncategorized / Flag for Review")

        time.sleep(1)

    return transactions


# ════════════════════════════════════════════════════════════════════════════
# AGENT 4 — INTELLIGENCE + FORECAST + HEALTH SCORE
# ════════════════════════════════════════════════════════════════════════════

GST_RATE_MAP = {
    "Software & Cloud Subscriptions":  0.18,
    "Office Rent & Utilities":         0.18,
    "Vendor Payments / Cost of Goods": 0.12,
}
GST_CONFIDENCE = {
    "Software & Cloud Subscriptions":  "HIGH",
    "Office Rent & Utilities":         "HIGH",
    "Vendor Payments / Cost of Goods": "LOW",
}
TALLY_LEDGER_MAP = {
    "Payroll & Salaries":               ("Salaries",            "Indirect Expenses"),
    "Software & Cloud Subscriptions":   ("Software Expenses",   "Indirect Expenses"),
    "Office Rent & Utilities":          ("Rent & Utilities",    "Indirect Expenses"),
    "Vendor Payments / Cost of Goods":  ("Purchase Accounts",   "Purchase Accounts"),
    "Tax Payments (GST, TDS)":          ("Duties & Taxes",      "Duties & Taxes"),
    "Bank Charges & Fees":              ("Bank Charges",        "Indirect Expenses"),
    "Revenue / Inward Payment":         ("Sales Accounts",      "Sales Accounts"),
    "Director Drawings / Personal":     ("Capital Account",     "Capital Account"),
    "Travel & Transport":               ("Travelling Expenses", "Indirect Expenses"),
    "ATM Withdrawal":                   ("Cash",                "Cash-in-Hand"),
    "Credit Card Payment":              ("Credit Card",         "Current Liabilities"),
    "Uncategorized / Flag for Review":  ("Suspense Account",    "Suspense Account"),
}


def generate_cash_flow_forecast(transactions: list) -> dict:
    """
    Groups spend by category per month.
    Forecasts next month using 3-month moving average.
    Returns { category: { months, history, forecast } }
    """
    monthly_spend = defaultdict(lambda: defaultdict(float))
    for t in transactions:
        cat   = t.get("CoA_Category", "Uncategorized / Flag for Review")
        debit = clean_amount(t.get("Debit", ""))
        mk    = month_key(t.get("Date", ""))
        if debit > 0 and mk != "0000-00":
            monthly_spend[cat][mk] += debit

    forecasts = {}
    for cat, month_data in monthly_spend.items():
        sorted_months = sorted(month_data.keys())
        history       = [round(month_data[m], 2) for m in sorted_months]
        window        = history[-3:]
        forecast      = round(sum(window) / len(window), 2)
        forecasts[cat] = {"months": sorted_months, "history": history, "forecast": forecast}

    return forecasts


def generate_health_score(summary: dict, transactions: list, forecasts: dict) -> dict:
    """
    5 dimensions x 20 points = 100 max.
    """
    score = 0
    breakdown = {}

    # 1. Net cash position
    net = summary["net_change"]
    if net > 0:
        pts, label = 20, "Cash flow positive"
    elif net > -(summary["total_outflow"] * 0.1):
        pts, label = 10, "Slightly negative"
    else:
        pts, label = 0, "Cash flow negative"
    score += pts
    breakdown["Cash Position"] = {"points": pts, "max": 20, "label": label}

    # 2. Anomaly rate
    total        = summary["total_transactions"] or 1
    anomaly_rate = summary["anomalies_detected"] / total
    if anomaly_rate == 0:
        pts, label = 20, "No anomalies detected"
    elif anomaly_rate < 0.05:
        pts, label = 10, f"{summary['anomalies_detected']} minor anomalies"
    else:
        pts, label = 0, f"{summary['anomalies_detected']} anomalies — review needed"
    score += pts
    breakdown["Anomaly Rate"] = {"points": pts, "max": 20, "label": label}

    # 3. GST compliance
    if summary["gst_eligible_spend"] > 0 and summary["estimated_itc"] > 0:
        pts, label = 20, f"ITC of Rs.{summary['estimated_itc']:,.0f} identified"
    elif summary["gst_eligible_spend"] > 0:
        pts, label = 10, "GST spend found, verify ITC"
    else:
        pts, label = 0, "No GST-eligible spend found"
    score += pts
    breakdown["GST Compliance"] = {"points": pts, "max": 20, "label": label}

    # 4. Data integrity (balance mismatches)
    mismatches = sum(1 for t in transactions if t.get("ValidationFlag") == "BALANCE_MISMATCH")
    if mismatches == 0:
        pts, label = 20, "All balances verified"
    elif mismatches <= 2:
        pts, label = 10, f"{mismatches} balance mismatches"
    else:
        pts, label = 0, f"{mismatches} mismatches — data quality risk"
    score += pts
    breakdown["Data Integrity"] = {"points": pts, "max": 20, "label": label}

    # 5. Spend stability (coefficient of variation across monthly spends)
    all_vals = [v for f in forecasts.values() for v in f["history"]]
    if len(all_vals) > 1:
        mean = sum(all_vals) / len(all_vals)
        std  = (sum((x - mean) ** 2 for x in all_vals) / len(all_vals)) ** 0.5
        cv   = std / mean if mean > 0 else 1
        if cv < 0.2:
            pts, label = 20, "Spend stable month-on-month"
        elif cv < 0.5:
            pts, label = 10, "Moderate spend variability"
        else:
            pts, label = 0, "High spend variability"
    else:
        pts, label = 10, "Single month — insufficient history"
    score += pts
    breakdown["Spend Stability"] = {"points": pts, "max": 20, "label": label}

    grade = "Excellent" if score >= 80 else "Good" if score >= 60 else "Fair" if score >= 40 else "Needs Attention"
    return {"score": score, "grade": grade, "breakdown": breakdown}


def generate_intelligence(transactions: list) -> dict:
    gst_categories    = list(GST_RATE_MAP.keys())
    spend_by_category = {}

    for t in transactions:
        cat   = t.get("CoA_Category", "Uncategorized / Flag for Review")
        debit = clean_amount(t.get("Debit", ""))
        if debit > 0:
            spend_by_category[cat] = round(spend_by_category.get(cat, 0) + debit, 2)

    gst_transactions   = []
    gst_eligible_total = 0.0
    estimated_itc      = 0.0

    for t in transactions:
        cat   = t.get("CoA_Category", "")
        debit = clean_amount(t.get("Debit", ""))
        if cat in gst_categories and debit > 0:
            rate  = GST_RATE_MAP[cat]
            itc   = round(debit * rate, 2)
            t["GST_Eligible"]   = True
            t["GST_Rate"]       = f"{int(rate*100)}%"
            t["GST_ITC_Est"]    = itc
            t["GST_Confidence"] = GST_CONFIDENCE[cat]
            gst_eligible_total += debit
            estimated_itc      += itc
            gst_transactions.append(t)
        else:
            t["GST_Eligible"]   = False
            t["GST_Rate"]       = ""
            t["GST_ITC_Est"]    = 0.0
            t["GST_Confidence"] = ""

    seen = {}
    for t in transactions:
        key = (t["Date"], t.get("Debit", ""), t.get("Credit", ""), t.get("Narration", "")[:30])
        if key in seen and (t.get("Debit") or t.get("Credit")):
            t["AnomalyFlag"] = "POSSIBLE_DUPLICATE"
        else:
            seen[key] = True
            t.setdefault("AnomalyFlag", "")

    debits = [clean_amount(t["Debit"]) for t in transactions if clean_amount(t.get("Debit", "")) > 0]
    if debits:
        avg = sum(debits) / len(debits)
        for t in transactions:
            d = clean_amount(t.get("Debit", ""))
            if d > avg * 5 and d > 10000 and not t.get("AnomalyFlag"):
                t["AnomalyFlag"] = "LARGE_DEBIT"

    total_debit   = sum(clean_amount(t.get("Debit",  "")) for t in transactions)
    total_credit  = sum(clean_amount(t.get("Credit", "")) for t in transactions)
    anomaly_count = sum(1 for t in transactions if t.get("AnomalyFlag"))

    summary = {
        "total_transactions": len(transactions),
        "total_outflow":      round(total_debit, 2),
        "total_inflow":       round(total_credit, 2),
        "net_change":         round(total_credit - total_debit, 2),
        "gst_eligible_spend": round(gst_eligible_total, 2),
        "estimated_itc":      round(estimated_itc, 2),
        "gst_transactions":   len(gst_transactions),
        "anomalies_detected": anomaly_count,
        "spend_by_category":  dict(sorted(spend_by_category.items(), key=lambda x: x[1], reverse=True)),
    }

    forecasts              = generate_cash_flow_forecast(transactions)
    health                 = generate_health_score(summary, transactions, forecasts)
    summary["forecasts"]         = forecasts
    summary["health_score"]      = health["score"]
    summary["health_grade"]      = health["grade"]
    summary["health_breakdown"]  = health["breakdown"]

    return summary


# ════════════════════════════════════════════════════════════════════════════
# AGENT 5 — CA CHAT
# ════════════════════════════════════════════════════════════════════════════

def chat_with_statement(transactions: list, summary: dict, user_question: str) -> str:
    api_key = os.getenv("GROQ_API_KEY")
    if not api_key:
        return "Error: GROQ_API_KEY not found in .env"

    client = Groq(api_key=api_key)

    forecast_lines = [
        f"  {cat}: forecast Rs.{f['forecast']:,.0f}/month"
        for cat, f in summary.get("forecasts", {}).items()
    ]

    context = f"""
You are a financial assistant for an Indian business CA workflow.

HEALTH SCORE: {summary.get('health_score', 'N/A')}/100 — {summary.get('health_grade', '')}

SUMMARY:
- Total Outflow:    Rs.{summary['total_outflow']:,.2f}
- Total Inflow:     Rs.{summary['total_inflow']:,.2f}
- Net Change:       Rs.{summary['net_change']:,.2f}
- GST ITC Estimate: Rs.{summary['estimated_itc']:,.2f} (needs GSTR-2B verification)
- Anomalies:        {summary['anomalies_detected']}

SPEND BY CATEGORY:
{json.dumps(summary['spend_by_category'], indent=2)}

NEXT MONTH FORECAST (3-month moving average):
{chr(10).join(forecast_lines) or 'Not available'}

TRANSACTIONS (first 50):
{json.dumps(transactions[:50], indent=2)}

Rules:
- Be specific with rupee amounts.
- Burn rate = total monthly outflow.
- Runway = current balance / monthly burn rate.
- Always note GST ITC needs GSTR-2B verification.
"""
    response = client.chat.completions.create(
        messages=[
            {"role": "system", "content": context},
            {"role": "user",   "content": user_question},
        ],
        model="llama-3.3-70b-versatile",
        temperature=0.1,
    )
    return response.choices[0].message.content


# ════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT — Transactions + Summary (with health score) + Forecast
# ════════════════════════════════════════════════════════════════════════════

def export_excel(transactions: list, summary: dict, path="finsight_output.xlsx"):
    PURPLE = "7353F6"
    wb = openpyxl.Workbook()

    hdr_fill  = PatternFill("solid", fgColor=PURPLE)
    hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri")
    red_fill  = PatternFill("solid", fgColor="FFEBEE")
    amb_fill  = PatternFill("solid", fgColor="FFF8E1")
    hlth_fill = PatternFill("solid", fgColor="EDE9FF")
    red_font  = Font(color="C62828", name="Calibri")
    grn_font  = Font(color="2E7D32", name="Calibri")
    pur_font  = Font(bold=True, color=PURPLE, name="Calibri")

    # ── Sheet 1: Transactions ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Transactions"
    headers    = ["Date", "Narration", "Debit", "Credit", "Balance",
                  "Category", "GST Eligible", "GST Rate", "ITC Estimate", "GST Confidence", "Flag"]
    col_widths = [12, 45, 14, 14, 14, 32, 13, 10, 14, 16, 20]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        ws.column_dimensions[c.column_letter].width = w

    for row, t in enumerate(transactions, 2):
        ws.cell(row=row, column=1,  value=t.get("Date", ""))
        ws.cell(row=row, column=2,  value=t.get("Narration", ""))
        dc = ws.cell(row=row, column=3, value=t.get("Debit", ""))
        if t.get("Debit"):   dc.font = red_font
        cc = ws.cell(row=row, column=4, value=t.get("Credit", ""))
        if t.get("Credit"):  cc.font = grn_font
        ws.cell(row=row, column=5,  value=t.get("Balance", ""))
        ws.cell(row=row, column=6,  value=t.get("CoA_Category", ""))
        ws.cell(row=row, column=7,  value="Yes" if t.get("GST_Eligible") else "")
        ws.cell(row=row, column=8,  value=t.get("GST_Rate", ""))
        itc = t.get("GST_ITC_Est", 0)
        ws.cell(row=row, column=9,  value=f"Rs.{itc:,.2f}" if itc else "")
        ws.cell(row=row, column=10, value=t.get("GST_Confidence", ""))
        ws.cell(row=row, column=11, value=t.get("AnomalyFlag", ""))
        if t.get("ValidationFlag") == "BALANCE_MISMATCH":
            for c in range(1, 12): ws.cell(row=row, column=c).fill = red_fill
        elif t.get("AnomalyFlag"):
            for c in range(1, 12): ws.cell(row=row, column=c).fill = amb_fill
    ws.freeze_panes = "A2"

    # ── Sheet 2: Summary ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions['A'].width = 38
    ws2.column_dimensions['B'].width = 24

    # health score — big and purple
    score = summary.get("health_score", 0)
    grade = summary.get("health_grade", "")
    for col, val in enumerate(["Financial Health Score", f"{score}/100 — {grade}"], 1):
        c = ws2.cell(row=1, column=col, value=val)
        c.fill = hlth_fill
        c.font = Font(bold=True, color=PURPLE, name="Calibri", size=13)

    summary_rows = [
        ("Total Transactions",              summary['total_transactions']),
        ("Total Outflow",                   f"Rs.{summary['total_outflow']:,.2f}"),
        ("Total Inflow",                    f"Rs.{summary['total_inflow']:,.2f}"),
        ("Net Change",                      f"Rs.{summary['net_change']:,.2f}"),
        ("GST Eligible Spend",              f"Rs.{summary['gst_eligible_spend']:,.2f}"),
        ("Estimated ITC (verify GSTR-2B)",  f"Rs.{summary['estimated_itc']:,.2f}"),
        ("Anomalies Detected",              summary['anomalies_detected']),
    ]
    for i, (label, value) in enumerate(summary_rows, 2):
        ws2.cell(row=i, column=1, value=label)
        ws2.cell(row=i, column=2, value=value)

    r = len(summary_rows) + 3
    ws2.cell(row=r, column=1, value="Health Score Breakdown").font = Font(bold=True, name="Calibri")
    r += 1
    for dim, info in summary.get("health_breakdown", {}).items():
        ws2.cell(row=r, column=1, value=dim)
        ws2.cell(row=r, column=2, value=f"{info['points']}/{info['max']} — {info['label']}")
        r += 1

    r += 1
    ws2.cell(row=r, column=1, value="Spend by Category").font = Font(bold=True, name="Calibri")
    r += 1
    for cat, amt in summary['spend_by_category'].items():
        ws2.cell(row=r, column=1, value=cat)
        ws2.cell(row=r, column=2, value=f"Rs.{amt:,.2f}")
        r += 1

    # ── Sheet 3: Cash Flow Forecast ───────────────────────────────────────
    ws3 = wb.create_sheet("Forecast")
    ws3.column_dimensions['A'].width = 38

    forecasts = summary.get("forecasts", {})
    if forecasts:
        all_months = sorted(set(m for f in forecasts.values() for m in f["months"]))

        ws3.cell(row=1, column=1, value="Category").fill = hdr_fill
        ws3.cell(row=1, column=1).font = hdr_font

        for col, m in enumerate(all_months, 2):
            ws3.column_dimensions[ws3.cell(1, col).column_letter].width = 16
            c = ws3.cell(row=1, column=col, value=m)
            c.fill = hdr_fill
            c.font = hdr_font

        fc_col = len(all_months) + 2
        ws3.column_dimensions[ws3.cell(1, fc_col).column_letter].width = 24
        fc = ws3.cell(row=1, column=fc_col, value="Next Month Forecast")
        fc.fill = hlth_fill
        fc.font = Font(bold=True, color=PURPLE, name="Calibri")

        for row, (cat, f) in enumerate(forecasts.items(), 2):
            ws3.cell(row=row, column=1, value=cat)
            month_data = dict(zip(f["months"], f["history"]))
            for col, m in enumerate(all_months, 2):
                val = month_data.get(m, None)
                ws3.cell(row=row, column=col, value=f"Rs.{val:,.0f}" if val is not None else "—")
            fcc = ws3.cell(row=row, column=fc_col, value=f"Rs.{f['forecast']:,.0f}")
            fcc.font = pur_font

    wb.save(path)
    print(f"Excel saved: {path}")


# ════════════════════════════════════════════════════════════════════════════
# TALLY XML EXPORT
# ════════════════════════════════════════════════════════════════════════════

def _sub(parent, tag, text=""):
    el = ET.SubElement(parent, tag)
    el.text = str(text)
    return el


def export_tally_xml(transactions: list, company_name: str = "My Business",
                     bank_ledger_name: str = "HDFC Bank Current A/c",
                     path: str = "finsight_tally.xml"):

    envelope = ET.Element("ENVELOPE")
    _sub(ET.SubElement(envelope, "HEADER"), "TALLYREQUEST", "Import Data")
    body = ET.SubElement(envelope, "BODY")

    # masters
    imp_m   = ET.SubElement(body, "IMPORTDATA")
    _sub(ET.SubElement(imp_m, "REQUESTDESC"), "REPORTNAME", "All Masters")
    tally_m = ET.SubElement(ET.SubElement(imp_m, "REQUESTDATA"), "TALLYMESSAGE", attrib={"xmlns:UDF": "TallyUDF"})

    unique_ledgers = {(bank_ledger_name, "Bank Accounts")}
    for t in transactions:
        cat = t.get("CoA_Category", "Uncategorized / Flag for Review")
        unique_ledgers.add(TALLY_LEDGER_MAP.get(cat, ("Suspense Account", "Suspense Account")))

    for name, group in sorted(unique_ledgers):
        lg = ET.SubElement(tally_m, "LEDGER", NAME=name, ACTION="Create")
        _sub(lg, "NAME", name)
        _sub(lg, "PARENT", group)
        _sub(lg, "TAXTYPE", "")

    # vouchers
    imp_v   = ET.SubElement(body, "IMPORTDATA")
    rd      = ET.SubElement(imp_v, "REQUESTDESC")
    _sub(rd, "REPORTNAME", "Vouchers")
    _sub(ET.SubElement(rd, "STATICVARIABLES"), "SVCURRENTCOMPANY", company_name) if False else \
        (lambda sv: _sub(sv, "SVCURRENTCOMPANY", company_name))(ET.SubElement(rd, "STATICVARIABLES"))
    tally_v = ET.SubElement(ET.SubElement(imp_v, "REQUESTDATA"), "TALLYMESSAGE", attrib={"xmlns:UDF": "TallyUDF"})

    skipped = 0
    for t in transactions:
        debit  = clean_amount(t.get("Debit",  ""))
        credit = clean_amount(t.get("Credit", ""))
        if debit == 0 and credit == 0:
            skipped += 1
            continue

        is_pay  = debit > 0
        amount  = debit if is_pay else credit
        vtype   = "Payment" if is_pay else "Receipt"
        cat     = t.get("CoA_Category", "Uncategorized / Flag for Review")
        ledger, _ = TALLY_LEDGER_MAP.get(cat, ("Suspense Account", "Suspense Account"))
        tdate   = (lambda dt: dt.strftime("%Y%m%d") if dt else "20000101")(parse_date(t.get("Date", "")))

        v = ET.SubElement(tally_v, "VOUCHER",
                          REMOTEID=f"FS-{tdate}-{int(amount)}", VCHTYPE=vtype, ACTION="Create")
        _sub(v, "DATE",            tdate)
        _sub(v, "VOUCHERTYPENAME", vtype)
        _sub(v, "NARRATION",       t.get("Narration", "")[:100])
        _sub(v, "VOUCHERNUMBER",   "")

        all_le = ET.SubElement(v, "ALLLEDGERENTRIES.LIST")
        if is_pay:
            le1 = ET.SubElement(all_le, "ALLLEDGERENTRIES.LIST")
            _sub(le1, "LEDGERNAME", ledger);  _sub(le1, "ISDEEMEDPOSITIVE", "Yes");  _sub(le1, "AMOUNT", f"-{amount:.2f}")
            le2 = ET.SubElement(all_le, "ALLLEDGERENTRIES.LIST")
            _sub(le2, "LEDGERNAME", bank_ledger_name); _sub(le2, "ISDEEMEDPOSITIVE", "No"); _sub(le2, "AMOUNT", f"{amount:.2f}")
        else:
            le1 = ET.SubElement(all_le, "ALLLEDGERENTRIES.LIST")
            _sub(le1, "LEDGERNAME", bank_ledger_name); _sub(le1, "ISDEEMEDPOSITIVE", "Yes"); _sub(le1, "AMOUNT", f"-{amount:.2f}")
            le2 = ET.SubElement(all_le, "ALLLEDGERENTRIES.LIST")
            _sub(le2, "LEDGERNAME", ledger); _sub(le2, "ISDEEMEDPOSITIVE", "No"); _sub(le2, "AMOUNT", f"{amount:.2f}")

    pretty = minidom.parseString(ET.tostring(envelope, encoding="unicode")).toprettyxml(indent="  ")
    with open(path, "w", encoding="utf-8") as f:
        f.write(pretty)
    print(f"Tally XML saved: {path}  ({len(transactions)-skipped} vouchers, {len(unique_ledgers)} ledgers)")
    return path


# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("Starting Agent 1 & 2 (Extractor + Validator)...")
    result = extract_hdfc_statement("sample_business_statement.pdf", password="")

    if result["status"] == "success":
        transactions = result["data"]
        mismatches   = sum(1 for t in transactions if t["ValidationFlag"] == "BALANCE_MISMATCH")
        print(f"SUCCESS — {len(transactions)} transactions | {mismatches} balance mismatches\n")

        print("Starting Agent 3 (Categorizer)...")
        categorized = categorize_with_groq(transactions)

        print("\nStarting Agent 4 (Intelligence + Forecast + Health Score)...")
        summary = generate_intelligence(categorized)

        print("\n" + "═" * 50)
        print(f"  FINANCIAL HEALTH SCORE: {summary['health_score']}/100 — {summary['health_grade']}")
        print("═" * 50)
        for dim, info in summary["health_breakdown"].items():
            bar = "█" * info["points"] + "░" * (info["max"] - info["points"])
            print(f"  {dim:<20} {bar}  {info['points']}/{info['max']}  {info['label']}")

        print(f"\n  Outflow:    Rs.{summary['total_outflow']:,.2f}")
        print(f"  Inflow:     Rs.{summary['total_inflow']:,.2f}")
        print(f"  Net:        Rs.{summary['net_change']:,.2f}")
        print(f"  ITC Est.:   Rs.{summary['estimated_itc']:,.2f}")
        print(f"  Anomalies:  {summary['anomalies_detected']}")

        print("\n  NEXT MONTH FORECAST (3-month moving avg):")
        for cat, f in summary["forecasts"].items():
            print(f"  {cat:<45} Rs.{f['forecast']:>12,.0f}")

        print("\nExporting Excel + Tally XML...")
        export_excel(categorized, summary)
        export_tally_xml(categorized,
                         company_name="My Business Pvt Ltd",
                         bank_ledger_name="HDFC Bank Current A/c")

        print("\nStarting Agent 5 (CA Chat)...")
        for q in ["What is my burn rate?",
                  "What is the biggest expense category?",
                  "How much GST input credit can I claim?",
                  "Are there any suspicious transactions?"]:
            print(f"\nQ: {q}")
            print(f"A: {chat_with_statement(categorized, summary, q)}")

    else:
        print(f"ERROR: {result['message']}")