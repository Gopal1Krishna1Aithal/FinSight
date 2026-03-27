import os
import sys
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')
import pandas as pd
from dotenv import load_dotenv

from core.extractors.hdfc_pdf import HDFCPDFExtractor
from core.processors.cleaner import HDFCDataCleaner
from core.processors.sanitizer import DataSanitizer
from core.ai_services.coa_mapper import CoAMapper, CONFIDENCE_THRESHOLD

from core.db.session import init_db
from core.db.operations import upsert_transactions
from core.ai_services.insights_generator import InsightsGenerator
import argparse


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

load_dotenv()

parser = argparse.ArgumentParser(description="Process a business bank statement PDF.")
parser.add_argument("pdf", nargs="?", default=os.path.join("data", "input", "pdfs"), help="Path to the PDF statement or folder of statements")
args = parser.parse_args()

PDF_PATH   = args.pdf
OUT_DIR    = os.path.join("data", "output")
EXCEL_PATH = os.path.join(OUT_DIR, "clean_statement.xlsx")
TALLY_PATH = os.path.join(OUT_DIR, "tally_import.csv")
INSIGHTS_PATH = os.path.join(OUT_DIR, "financial_insights.md")


# ---------------------------------------------------------------------------
# Step 3 — Mathematical Validator
# ---------------------------------------------------------------------------

def validate_balances(df: pd.DataFrame) -> bool:
    """
    Walks every row and verifies:
        previous_balance - debit + credit  ≈  current_balance  (±₹0.01)

    Opening balance is back-calculated from row 0:
        opening = balance[0] + debit[0] - credit[0]
    """
    TOLERANCE    = 0.01
    prev_balance = df.iloc[0]["Balance"] + df.iloc[0]["Debit"] - df.iloc[0]["Credit"]

    for idx, row in df.iterrows():
        expected = prev_balance - row["Debit"] + row["Credit"]
        actual   = row["Balance"]
        if abs(expected - actual) > TOLERANCE:
            print(
                f"\n      [VALIDATOR] ❌  Mismatch on row {idx} "
                f"({row['Date'].strftime('%d/%m/%Y')}):\n"
                f"        Expected : ₹{expected:,.2f}\n"
                f"        Actual   : ₹{actual:,.2f}"
            )
            return False
        prev_balance = actual

    return True


# ---------------------------------------------------------------------------
# Step 5 — Output writers
# ---------------------------------------------------------------------------

def _save_excel(df: pd.DataFrame, path: str) -> None:
    """
    Writes a two-sheet Excel workbook:

    Sheet 1 — "Transactions"
        All 173 rows with columns:
        Date | Narration | Clean_Description | CoA_Category |
        Confidence_Score | Reasoning | Review_Required | Debit | Credit | Balance

        Review_Required = TRUE when Confidence_Score < threshold OR category is Uncategorized.
        CA workflow: filter Review_Required = TRUE → fix only those rows, ignore the rest.

    Sheet 2 — "Summary"
        Financial totals by category + overall inflow/outflow/net.
        This is what a CA looks at first before drilling into transactions.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("      [!] openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    # ── Build the transactions DataFrame ──────────────────────────────
    out = df[[
        "Date", "Narration", "Clean_Description", "CoA_Category",
        "Confidence_Score", "Reasoning", "Debit", "Credit", "Balance"
    ]].copy()

    out["Date"] = out["Date"].dt.strftime("%d/%m/%Y")

    # Review_Required: True if confidence low OR category is Uncategorized
    out["Review_Required"] = (
        (df["Confidence_Score"] < CONFIDENCE_THRESHOLD) |
        (df["CoA_Category"] == "Uncategorized")
    )

    # Reorder so Review_Required is visible right after CoA columns
    out = out[[
        "Date", "Narration", "Clean_Description", "CoA_Category",
        "Confidence_Score", "Reasoning", "Review_Required",
        "Debit", "Credit", "Balance"
    ]]

    # ── Build the summary DataFrame ───────────────────────────────────
    summary_rows = []

    # Per-category totals
    for cat in sorted(df["CoA_Category"].unique()):
        cat_df   = df[df["CoA_Category"] == cat]
        total_dr = cat_df["Debit"].sum()
        total_cr = cat_df["Credit"].sum()
        count    = len(cat_df)
        summary_rows.append({
            "Category":    cat,
            "Txn Count":   count,
            "Total Debit": round(total_dr, 2),
            "Total Credit":round(total_cr, 2),
            "Net":         round(total_cr - total_dr, 2),
        })

    summary_df = pd.DataFrame(summary_rows)

    # Overall totals row
    overall = {
        "Category":     "── TOTAL ──",
        "Txn Count":    len(df),
        "Total Debit":  round(df["Debit"].sum(), 2),
        "Total Credit": round(df["Credit"].sum(), 2),
        "Net":          round(df["Credit"].sum() - df["Debit"].sum(), 2),
    }

    # Opening / closing balance
    opening_balance = df.iloc[0]["Balance"] + df.iloc[0]["Debit"] - df.iloc[0]["Credit"]
    closing_balance = df.iloc[-1]["Balance"]

    # Flagging stats
    review_count = out["Review_Required"].sum()

    # ── Write to Excel ────────────────────────────────────────────────
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # Sheet 1 — Transactions
        out.to_excel(writer, sheet_name="Transactions", index=False)
        ws_txn = writer.sheets["Transactions"]

        # Freeze header row
        ws_txn.freeze_panes = "A2"

        # Column widths
        col_widths = {
            "A": 12,   # Date
            "B": 45,   # Narration
            "C": 28,   # Clean_Description
            "D": 26,   # CoA_Category
            "E": 16,   # Confidence_Score
            "F": 40,   # Reasoning
            "G": 16,   # Review_Required
            "H": 14,   # Debit
            "I": 14,   # Credit
            "J": 14,   # Balance
        }
        for col, width in col_widths.items():
            ws_txn.column_dimensions[col].width = width

        # Style header row
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        for cell in ws_txn[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws_txn.row_dimensions[1].height = 30

        # Highlight Review_Required = True rows in yellow
        review_fill  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        flag_col_idx = out.columns.get_loc("Review_Required") + 1   # 1-indexed

        for row_idx, row_val in enumerate(out["Review_Required"], start=2):
            if row_val:
                for col_idx in range(1, len(out.columns) + 1):
                    ws_txn.cell(row=row_idx, column=col_idx).fill = review_fill
                # Also make the Review_Required cell itself red-bold
                flag_cell            = ws_txn.cell(row=row_idx, column=flag_col_idx)
                flag_cell.font       = Font(bold=True, color="C00000")
                flag_cell.alignment  = Alignment(horizontal="center")

        # Sheet 2 — Summary
        summary_df.to_excel(writer, sheet_name="Summary", index=False, startrow=6)
        ws_sum = writer.sheets["Summary"]
        ws_sum.column_dimensions["A"].width = 28
        ws_sum.column_dimensions["B"].width = 12
        ws_sum.column_dimensions["C"].width = 16
        ws_sum.column_dimensions["D"].width = 16
        ws_sum.column_dimensions["E"].width = 14

        # Header block at the top of Summary sheet
        title_font  = Font(bold=True, size=13, color="1F4E79")
        label_font  = Font(bold=True, size=10)
        value_font  = Font(size=10)
        green_font  = Font(bold=True, color="375623", size=10)
        red_font    = Font(bold=True, color="C00000", size=10)

        ws_sum["A1"] = "FINANCIAL SUMMARY"
        ws_sum["A1"].font = title_font

        kv_rows = [
            ("Opening Balance",  f"₹{opening_balance:,.2f}"),
            ("Closing Balance",  f"₹{closing_balance:,.2f}"),
            ("Total Inflow",     f"₹{df['Credit'].sum():,.2f}"),
            ("Total Outflow",    f"₹{df['Debit'].sum():,.2f}"),
            ("Net Cash Flow",    f"₹{df['Credit'].sum() - df['Debit'].sum():,.2f}"),
            ("Transactions",     str(len(df))),
            ("Flagged for Review", str(int(review_count))),
        ]

        for i, (label, value) in enumerate(kv_rows, start=2):
            ws_sum.cell(row=i, column=1).value = label
            ws_sum.cell(row=i, column=1).font  = label_font
            ws_sum.cell(row=i, column=2).value = value
            # Colour net cash flow green/red
            if label == "Net Cash Flow":
                net_val = df["Credit"].sum() - df["Debit"].sum()
                ws_sum.cell(row=i, column=2).font = green_font if net_val >= 0 else red_font
            elif label == "Flagged for Review" and review_count > 0:
                ws_sum.cell(row=i, column=2).font = red_font
            else:
                ws_sum.cell(row=i, column=2).font = value_font

        # Style the category breakdown table header (row 7)
        for cell in ws_sum[7]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center")

        # Total row at the bottom of the category table
        total_row_idx = 7 + len(summary_df) + 1
        total_fill    = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        total_data    = [
            overall["Category"],
            overall["Txn Count"],
            overall["Total Debit"],
            overall["Total Credit"],
            overall["Net"],
        ]
        for col_idx, val in enumerate(total_data, start=1):
            cell            = ws_sum.cell(row=total_row_idx, column=col_idx)
            cell.value      = val
            cell.fill       = total_fill
            cell.font       = Font(bold=True, size=10)
            cell.alignment  = Alignment(horizontal="center")

    print(f"      ✅  Excel  → {path}  ({int(review_count)} rows flagged for review)")


def _save_tally_csv(df: pd.DataFrame, path: str) -> None:
    """
    Tally-ready CSV — clean data only, no review flags.
    Columns: Date | Voucher_Type | Ledger_Name | CoA_Category | Amount
    """
    tally = df[["Date", "Clean_Description", "CoA_Category", "Debit", "Credit"]].copy()
    tally["Date"]         = df["Date"].dt.strftime("%d/%m/%Y")
    tally["Voucher_Type"] = tally.apply(lambda r: "Payment" if r["Debit"] > 0 else "Receipt", axis=1)
    tally["Amount"]       = tally.apply(lambda r: r["Debit"] if r["Debit"] > 0 else r["Credit"], axis=1)
    tally = tally.rename(columns={"Clean_Description": "Ledger_Name"})
    tally = tally[["Date", "Voucher_Type", "Ledger_Name", "CoA_Category", "Amount"]]
    tally.to_csv(path, index=False)
    print(f"      ✅  Tally  → {path}")


# ---------------------------------------------------------------------------
# Pipeline
# ---------------------------------------------------------------------------

def run_pipeline() -> None:
    global PDF_PATH
    os.makedirs(OUT_DIR, exist_ok=True)

    print("\n[0/6] Initializing database...")
    init_db()

    # ── Step 1: Extract ──────────────────────────────────────────────────
    print(f"\n[1/6] Extracting raw transactions from '{PDF_PATH}'...")
    if not os.path.exists(PDF_PATH):
        print(f"      [!] Path not found at '{PDF_PATH}'. Aborting.")
        sys.exit(1)

    raw_data = []
    if os.path.isdir(PDF_PATH):
        # If the default pdfs folder is explicitly empty, check the images folder as a convenience fallback
        if PDF_PATH == os.path.join("data", "input", "pdfs") and not os.listdir(PDF_PATH):
             fallback_path = os.path.join("data", "input", "images")
             if os.path.exists(fallback_path) and any(f.lower().endswith(('.jpg', '.jpeg', '.png', '.heic')) for f in os.listdir(fallback_path)):
                 print(f"      [!] '{PDF_PATH}' is empty. Falling back to '{fallback_path}'.")
                 PDF_PATH = fallback_path

        # Folder Mode (Image or PDF)
        valid_img_exts = {".jpg", ".jpeg", ".png", ".heic"}
        image_files = [
            os.path.join(PDF_PATH, f) for f in os.listdir(PDF_PATH)
            if os.path.splitext(f.lower())[1] in valid_img_exts
        ]
        pdf_files = [
            os.path.join(PDF_PATH, f) for f in os.listdir(PDF_PATH)
            if f.lower().endswith('.pdf')
        ]
        
        if image_files:
            print(f"      [OCR] Found {len(image_files)} image files in {PDF_PATH}.")
            from core.extractors.image_ocr import ImageOCRExtractor
            extractor = ImageOCRExtractor(image_paths=image_files)
            raw_data = extractor.extract() or []
        elif pdf_files:
            print(f"      [PDF] Found {len(pdf_files)} PDF files in {PDF_PATH}. Processing sequentially...")
            for pf in sorted(pdf_files):
                print(f"            → Extracting {os.path.basename(pf)}...")
                ext_data = HDFCPDFExtractor(pf).extract()
                if ext_data:
                    raw_data.extend(ext_data)
        else:
            print(f"      [!] No valid image or PDF files found in '{PDF_PATH}'. Aborting.")
            sys.exit(1)
    else:
        # Standard Single File Extractor
        if PDF_PATH.lower().endswith('.pdf'):
            print(f"      [PDF] Extracting single file {os.path.basename(PDF_PATH)}...")
            raw_data = HDFCPDFExtractor(PDF_PATH).extract() or []
        else:
            print(f"      [!] File {PDF_PATH} is not a PDF. Aborting.")
            sys.exit(1)

    if not raw_data:
        print("      [!] Extraction returned no data. Aborting.")
        sys.exit(1)
    print(f"      → {len(raw_data)} rows extracted.")

    # ── Step 2: Clean ────────────────────────────────────────────────────
    print("\n[2/6] Cleaning narrations and coercing numbers...")
    clean_df = HDFCDataCleaner(raw_data).clean()
    print(f"      → {len(clean_df)} rows | null dates: {clean_df['Date'].isna().sum()}")

    # ── Step 3: Validate ─────────────────────────────────────────────────
    print("\n[3/6] Validating balance integrity...")
    if not validate_balances(clean_df):
        print("\n      [!] Validation FAILED — fix extraction before proceeding.")
        sys.exit(1)
    print(f"      → All {len(clean_df)} balances verified ✅")

    # ── Step 4: Scrub PII ────────────────────────────────────────────────
    print("\n[4/6] Scrubbing PII and building Clean_Description...")
    safe_df = DataSanitizer(clean_df).scrub_pii()
    sample  = safe_df[["Narration", "Clean_Description"]].drop_duplicates().head(5)
    for _, row in sample.iterrows():
        print(f"      {row['Narration'][:42]:<42}  →  {row['Clean_Description']}")

    # ── Step 4.5: CoA Categorisation + Confidence Scoring ────────────────
    print("\n[4.5/6] Categorising via Groq LLM (with confidence scoring)...")
    groq_api_key = os.getenv("GROQ_API_KEY")

    if not groq_api_key:
        print(
            "      [!] GROQ_API_KEY not set in .env — skipping LLM categorisation.\n"
            "          All rows will be marked Uncategorized / Review_Required = True."
        )
        safe_df["CoA_Category"]     = "Uncategorized"
        safe_df["Confidence_Score"] = 0
        safe_df["Reasoning"]        = "GROQ_API_KEY not configured."
    else:
        mapper  = CoAMapper(api_key=groq_api_key)
        safe_df = mapper.map(safe_df)

    # ── Step 5: Save outputs ─────────────────────────────────────────────
    print("\n[5/6] Writing output files and updating database...")
    _save_excel(safe_df, EXCEL_PATH)
    _save_tally_csv(safe_df, TALLY_PATH)

    new_rows = upsert_transactions(safe_df)
    print(f"      ✅  Database → Inserted {new_rows} new transactions.")

    # ── Step 6: Generate Insights ────────────────────────────────────────
    print("\n[6/6] Generating financial insights from full history...")
    try:
        generator = InsightsGenerator()
        generator.generate_insights(INSIGHTS_PATH)
    except Exception as e:
        print(f"      [Insights] Initialization failed: {e}")

    # ── Terminal summary ─────────────────────────────────────────────────
    review_count = int(
        ((safe_df["Confidence_Score"] < CONFIDENCE_THRESHOLD) |
         (safe_df["CoA_Category"] == "Uncategorized")).sum()
    )

    print(f"\n{'─' * 58}")
    print(f"  Pipeline complete — {len(safe_df)} transactions processed")
    print(f"  Opening balance : ₹{clean_df.iloc[0]['Balance'] + clean_df.iloc[0]['Debit'] - clean_df.iloc[0]['Credit']:>12,.2f}")
    print(f"  Total inflow    : ₹{clean_df['Credit'].sum():>12,.2f}")
    print(f"  Total outflow   : ₹{clean_df['Debit'].sum():>12,.2f}")
    net = clean_df['Credit'].sum() - clean_df['Debit'].sum()
    print(f"  Net cash flow   : ₹{net:>12,.2f}  {'▲' if net >= 0 else '▼'}")
    print(f"  Closing balance : ₹{clean_df.iloc[-1]['Balance']:>12,.2f}")
    print(f"  Flagged rows    : {review_count}  (Confidence < {CONFIDENCE_THRESHOLD}% or Uncategorized)")
    print(f"\n  Category breakdown:")
    for cat, count in safe_df["CoA_Category"].value_counts().items():
        bar = "█" * min(count, 30)
        print(f"    {cat:<28} {count:>4}  {bar}")
    print(f"\n  Output : {os.path.abspath(OUT_DIR)}")
    print(f"{'─' * 58}\n")


if __name__ == "__main__":
    run_pipeline()